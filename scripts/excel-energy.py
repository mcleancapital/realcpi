import os
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import requests
from bs4 import BeautifulSoup

# === CONFIG ===
EXCEL_FILE_PATH = './data/energy.xlsx'
SHEET_NAME = "Data"
URL = "https://ycharts.com/indicators/us_primary_energy_consumption"
HEADERS = {
    "User-Agent": "Mozilla/5.0",
    "Accept-Language": "en-US,en;q=0.9",
}


def fetch_recent_energy_data(url):
    """Fetch the most recent energy consumption data from YCharts."""
    try:
        response = requests.get(url, headers=HEADERS)
        if response.status_code != 200:
            raise Exception(f"Request failed with status code {response.status_code}")

        soup = BeautifulSoup(response.content, "html.parser")
        element = soup.find("div", class_="key-stat-title")
        if not element:
            raise Exception("Could not find 'key-stat-title' on the page.")

        text = element.get_text(strip=True)
        parts = text.split(maxsplit=2)
        if len(parts) < 3:
            raise Exception("Unexpected data format from YCharts.")

        value = float(parts[0].replace('Q', ''))  # Example: 97.23Q
        date = datetime.strptime(parts[2].replace("for ", ""), "%b %Y").strftime("%Y-%m-%d")

        print(f"✔️  Fetched energy value: {value}Q BTU for {date}")
        return date, value
    except Exception as e:
        print(f"❌ Error fetching energy data: {e}")
        return None, None


def update_excel_with_energy(file_path, recent_date, recent_value):
    """Update Excel with new energy data and compute % change vs last year using Python."""
    try:
        if not os.path.exists(file_path):
            print(f"❌ File not found: {file_path}")
            return

        wb = load_workbook(file_path)
        if SHEET_NAME not in wb.sheetnames:
            print(f"❌ Sheet '{SHEET_NAME}' not found.")
            return
        ws = wb[SHEET_NAME]

        # Read existing data
        existing_data = []
        for row in range(2, ws.max_row + 1):
            date_val = ws.cell(row=row, column=1).value
            value_val = ws.cell(row=row, column=2).value
            existing_data.append((date_val, value_val))

        # Check for duplicate date
        if existing_data:
            most_recent_in_excel = existing_data[0][0]
            if pd.to_datetime(most_recent_in_excel).strftime("%Y-%m-%d") == recent_date:
                print(f"ℹ️  {recent_date} already exists — skipping insert.")
                return

        # Insert new row at top
        ws.insert_rows(2)
        ws.cell(row=2, column=1, value=recent_date)
        ws.cell(row=2, column=2, value=recent_value)

        # Rebuild data list with new entry on top
        existing_data.insert(0, (recent_date, recent_value))

        # Calculate % change vs last year using Python
        for i in range(len(existing_data)):
            row_excel = i + 2  # account for header row
            if i + 12 < len(existing_data):
                curr_val = existing_data[i][1]
                prev_val = existing_data[i + 12][1]
                if curr_val is not None and prev_val not in [0, None]:
                    change = (curr_val / prev_val - 1) * 100
                    ws.cell(row=row_excel, column=3, value=round(change, 1))
                else:
                    ws.cell(row=row_excel, column=3, value=None)
            else:
                ws.cell(row=row_excel, column=3, value=None)

        # Save final workbook
        wb.save(file_path)
        print("✅ Excel file updated successfully with calculated values.")

    except Exception as e:
        print(f"❌ Error updating Excel: {e}")


if __name__ == "__main__":
    date, value = fetch_recent_energy_data(URL)
    if date and value:
        update_excel_with_energy(EXCEL_FILE_PATH, date, value)

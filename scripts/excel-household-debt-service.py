import requests
from openpyxl import Workbook, load_workbook
from datetime import datetime

# FRED API Setup
API_KEY = "dc0a72da11fb0ffc6f4dc8d31e7afbb5"  # Replace with your own key
FRED_BASE = "https://api.stlouisfed.org/fred/series/observations"
SERIES_ID = "TDSP"  # Household Debt Service Payments as a % of Disposable Personal Income

# Output Excel file
EXCEL_FILE = "./data/household-debt-service.xlsx"

# Fetch data from FRED
def get_fred_quarterly_data(series_id, api_key):
    params = {
        "series_id": series_id,
        "api_key": api_key,
        "file_type": "json",
        "frequency": "q"  # Quarterly frequency
    }
    response = requests.get(FRED_BASE, params=params)
    response.raise_for_status()
    data = response.json()
    if "observations" in data:
        return {
            obs["date"]: float(obs["value"])
            for obs in data["observations"]
            if obs["value"] != "."
        }
    else:
        raise Exception(f"No data found for series {series_id}")

# Main logic
try:
    tdsp_data = get_fred_quarterly_data(SERIES_ID, API_KEY)

    # Get the most recent quarter
    latest_date = max(tdsp_data.keys())
    latest_value = tdsp_data[latest_date]

    # Load or create Excel file
    try:
        wb = load_workbook(EXCEL_FILE)
        sheet = wb.active
    except FileNotFoundError:
        wb = Workbook()
        sheet = wb.active
        sheet.append(["Quarter End", "Debt Service (% Disposable Income)", "% Change vs B14"])

    # Get existing dates to avoid duplicates
    existing_dates = [str(sheet.cell(row=i, column=1).value) for i in range(2, sheet.max_row + 1)]

    if latest_date not in existing_dates:
        print(f"Inserting new row for {latest_date}...")

        # Insert new row at the top (after header)
        sheet.insert_rows(2)
        sheet.cell(row=2, column=1, value=latest_date)
        sheet.cell(row=2, column=2, value=latest_value)

        # Get B14 value (row 14, column 2)
        b14_value = sheet.cell(row=14, column=2).value
        if isinstance(b14_value, (int, float)) and b14_value != 0:
            percent_change = (latest_value / b14_value - 1) * 100
            sheet.cell(row=2, column=3, value=round(percent_change, 2))
        else:
            sheet.cell(row=2, column=3, value=None)  # Or "N/A" if you prefer

    else:
        print(f"{latest_date} already exists — skipping insertion.")

    # Save the workbook
    wb.save(EXCEL_FILE)
    print(f"{latest_date}: {latest_value:.2f}% saved to {EXCEL_FILE}")

except Exception as e:
    print(f"Error: {e}")

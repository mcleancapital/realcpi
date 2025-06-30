import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter
import requests
from bs4 import BeautifulSoup
from datetime import datetime

# File path to the Excel file
EXCEL_FILE_PATH = './data/population.xlsx'

# URL and headers
URL = "https://ycharts.com/indicators/us_total_population"
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36",
    "Accept-Language": "en-US,en;q=0.9",
}

def fetch_recent_population_data(url):
    """Fetch the most recent US population data from YCharts."""
    try:
        response = requests.get(url, headers=HEADERS)
        if response.status_code != 200:
            raise Exception(f"Request failed with status code {response.status_code}")

        soup = BeautifulSoup(response.content, "html.parser")
        element = soup.find("div", class_="key-stat-title")
        if not element:
            raise Exception("Failed to find population data element.")

        text = element.get_text(strip=True)
        parts = text.split(maxsplit=2)
        if len(parts) < 2:
            raise Exception("Unexpected format for population data.")

        value = float(parts[0].replace('M', '').replace(',', ''))  # Value in millions
        date_str = parts[2].replace("for ", "")
        date = datetime.strptime(date_str, "%b %Y").strftime("%Y-%m-%d")

        print(f"Fetched population data - Value: {value}, Date: {date}")
        return date, value
    except Exception as e:
        print(f"Error fetching population data: {e}")
        return None, None

def update_excel(file_path, recent_date, recent_value):
    """Update the Excel file with the new population data and a formula in column C."""
    try:
        if not os.path.exists(file_path):
            print(f"File not found: {file_path}")
            return

        wb = load_workbook(file_path)
        sheet_name = "Data"
        if sheet_name not in wb.sheetnames:
            print(f"'{sheet_name}' sheet not found in the workbook.")
            return
        ws = wb[sheet_name]

        # Check if the date already exists
        most_recent_date_in_excel = ws.cell(row=2, column=1).value
        if most_recent_date_in_excel:
            most_recent_date_in_excel = pd.to_datetime(most_recent_date_in_excel).strftime("%Y-%m-%d")

        if most_recent_date_in_excel == recent_date:
            print(f"Recent date {recent_date} already exists. No update.")
            return

        # Insert new row at row 2
        ws.insert_rows(2)
        ws.cell(row=2, column=1, value=recent_date)   # Column A
        ws.cell(row=2, column=2, value=recent_value)  # Column B

        # Write formula in Column C
        formula = "=(B2/B14-1)*100"
        ws.cell(row=2, column=3, value=f"={formula}")  # Column C

        print(f"Added row with formula in C2: {formula}")

        wb.save(file_path)
        print(f"Excel file updated successfully: {file_path}")
    except Exception as e:
        print(f"Error updating Excel file: {e}")

if __name__ == "__main__":
    date, value = fetch_recent_population_data(URL)
    if date and value:
        update_excel(EXCEL_FILE_PATH, date, value)

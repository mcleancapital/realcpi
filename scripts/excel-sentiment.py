import os
import pandas as pd
from openpyxl import load_workbook
import requests
from bs4 import BeautifulSoup
from datetime import datetime

# File path to the Excel file
EXCEL_FILE_PATH = './data/consumer-sentiment.xlsx'

# URL and headers
URL = "https://ycharts.com/indicators/us_consumer_sentiment_index"
HEADERS = {
    "User-Agent": "Mozilla/5.0",
    "Accept-Language": "en-US,en;q=0.9",
}

def fetch_recent_consumer_sentiment_data(url):
    """Fetch the most recent U.S. consumer sentiment index data from YCharts."""
    try:
        response = requests.get(url, headers=HEADERS)
        if response.status_code != 200:
            raise Exception(f"Request failed with status code {response.status_code}")

        soup = BeautifulSoup(response.content, "html.parser")
        element = soup.find("div", class_="key-stat-title")
        if not element:
            raise Exception("Failed to find the element with class 'key-stat-title'.")

        text = element.get_text(strip=True)
        parts = text.split(maxsplit=2)
        if len(parts) < 3:
            raise Exception("Unexpected format for consumer sentiment data.")

        value = float(parts[0].replace(',', ''))
        date_str = parts[2].replace("for ", "")
        date = datetime.strptime(date_str.strip(), "%b %Y").strftime("%Y-%m-%d")

        print(f"Fetched Consumer Sentiment - Value: {value}, Date: {date}")
        return date, value
    except Exception as e:
        print(f"Error fetching consumer sentiment data: {e}")
        return None, None

def update_excel(file_path, recent_date, recent_value):
    """Update the Excel file with the most recent consumer sentiment data and compute YoY % in column C."""
    try:
        if not os.path.exists(file_path):
            print(f"File not found: {file_path}")
            return

        wb = load_workbook(file_path)
        sheet_name = "Data"
        if sheet_name not in wb.sheetnames:
            print(f"'{sheet_name}' sheet not found in the workbook.")
            return
        ws = wb[sheet_name]

        # Check if date already exists
        most_recent_date_in_excel = ws.cell(row=2, column=1).value
        if most_recent_date_in_excel:
            most_recent_date_in_excel = pd.to_datetime(most_recent_date_in_excel).strftime("%Y-%m-%d")
        if most_recent_date_in_excel == recent_date:
            print(f"Recent date {recent_date} already exists in the Excel file. No update needed.")
            return

        # Insert a new row at the top
        ws.insert_rows(2)
        ws.cell(row=2, column=1, value=recent_date)
        ws.cell(row=2, column=2, value=recent_value)

        # Build list of values
        values = []
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row=row, column=2).value
            values.append(val)

        # Compute YoY % in column C using Python
        for i in range(len(values)):
            if i + 12 < len(values):
                prev = values[i + 12]
                curr = values[i]
                if prev and prev != 0:
                    yoy = ((curr / prev) - 1) * 100
                    ws.cell(row=i + 2, column=3, value=round(yoy, 2))
                else:
                    ws.cell(row=i + 2, column=3, value="N/A")
            else:
                ws.cell(row=i + 2, column=3, value="")

        wb.save(file_path)
        print(f"Excel file updated successfully with consumer sentiment data and Python-computed YoY %.")
    except Exception as e:
        print(f"Error updating Excel file: {e}")

if __name__ == "__main__":
    recent_date, recent_value = fetch_recent_consumer_sentiment_data(URL)
    if recent_date and recent_value:
        update_excel(EXCEL_FILE_PATH, recent_date, recent_value)

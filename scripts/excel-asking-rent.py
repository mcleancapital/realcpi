import os
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils.cell import get_column_letter
import requests
from bs4 import BeautifulSoup
from datetime import datetime

# File path to the Excel file
EXCEL_FILE_PATH = './data/asking-rent.xlsx'

# URL and headers
URL = "https://ycharts.com/indicators/us_median_asking_rent"
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36",
    "Accept-Language": "en-US,en;q=0.9",
}

def ensure_excel_exists(path):
    """Create Excel file and Data sheet if missing."""
    if not os.path.exists(path):
        print(f"{path} not found. Creating new file.")
        os.makedirs(os.path.dirname(path), exist_ok=True)
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        ws.append(["Period", "Asking Rent (USD)", "YoY (%)"])
        wb.save(path)
        print(f"Created file with headers: {path}")

def fetch_recent_rent_data(url):
    """Fetch the most recent asking rent data from the YCharts page."""
    try:
        response = requests.get(url, headers=HEADERS)
        if response.status_code != 200:
            raise Exception(f"Request failed with status code {response.status_code}")

        soup = BeautifulSoup(response.content, "html.parser")
        rent_element = soup.find("div", class_="key-stat-title")
        if not rent_element:
            raise Exception("Failed to find the element with class 'key-stat-title'.")

        rent_text = rent_element.get_text(strip=True)
        parts = rent_text.replace("USD", "").split("for")
        if len(parts) != 2:
            raise Exception("Unexpected format for rent data.")

        value = float(parts[0].strip().replace(",", ""))
        quarter_str = parts[1].strip()  # e.g., "Q1 2025"

        # Convert to a more sortable format: "2025-Q1"
        formatted_period = f"{quarter_str[-4:]}-{quarter_str[:2]}"

        print(f"Fetched Rent data - Value: {value}, Period: {formatted_period}")
        return formatted_period, value
    except Exception as e:
        print(f"Error fetching rent data: {e}")
        return None, None

def update_excel(file_path, recent_period, recent_value):
    """Update the Excel file with the most recent asking rent data."""
    try:
        wb = load_workbook(file_path)
        sheet_name = "Data"
        if sheet_name not in wb.sheetnames:
            print(f"'{sheet_name}' sheet not found. Creating it.")
            ws = wb.create_sheet(title=sheet_name)
            ws.append(["Period", "Asking Rent (USD)", "YoY (%)"])
        else:
            ws = wb[sheet_name]

        # Check if recent period already exists in A2
        most_recent_period = ws.cell(row=2, column=1).value
        if most_recent_period == recent_period:
            print(f"Recent period {recent_period} already exists. No update needed.")
            return

        # Insert a new row at the top (after headers)
        ws.insert_rows(2)
        ws.cell(row=2, column=1, value=recent_period)
        ws.cell(row=2, column=2, value=recent_value)

        # Update column C: YoY = row / (row+3) - 1
        for row in range(2, ws.max_row + 1):
            try:
                curr = ws.cell(row=row, column=2).value
                prev = ws.cell(row=row + 3, column=2).value
                if isinstance(curr, (int, float)) and isinstance(prev, (int, float)) and prev != 0:
                    yoy = (curr / prev - 1) * 100
                    ws.cell(row=row, column=3, value=round(yoy, 2))
                else:
                    ws.cell(row=row, column=3, value=None)
            except:
                ws.cell(row=row, column=3, value=None)

        wb.save(file_path)
        print(f"Excel file updated successfully: {file_path}")
    except Exception as e:
        print(f"Error updating Excel file: {e}")

if __name__ == "__main__":
    ensure_excel_exists(EXCEL_FILE_PATH)
    recent_period, recent_value = fetch_recent_rent_data(URL)
    if recent_period and recent_value:
        update_excel(EXCEL_FILE_PATH, recent_period, recent_value)

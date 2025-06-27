import os
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import requests
from bs4 import BeautifulSoup

# File path to the Excel file
EXCEL_FILE_PATH = './data/existing-home-sales.xlsx'

# URL and headers
URL = "https://ycharts.com/indicators/us_existing_home_sales"
HEADERS = {
    "User-Agent": "Mozilla/5.0",
    "Accept-Language": "en-US,en;q=0.9",
}

def fetch_recent_sales_data(url):
    """Fetch the most recent Existing Home Sales data from YCharts."""
    try:
        response = requests.get(url, headers=HEADERS)
        if response.status_code != 200:
            raise Exception(f"Request failed with status code {response.status_code}")

        soup = BeautifulSoup(response.content, "html.parser")
        sales_element = soup.find("div", class_="key-stat-title")
        if not sales_element:
            raise Exception("Failed to find the element with class 'key-stat-title'.")

        sales_text = sales_element.get_text(strip=True)
        parts = sales_text.split(maxsplit=2)
        if len(parts) < 3:
            raise Exception("Unexpected format for sales data.")

        recent_value = float(parts[0].replace('M', ''))  # e.g. "4.1M"
        recent_date_str = parts[2].replace("for ", "")   # e.g. "May 2024"
        recent_date = datetime.strptime(recent_date_str, "%b %Y").strftime("%Y-%m-%d")

        print(f"Fetched Existing Home Sales – Value: {recent_value}, Date: {recent_date}")
        return recent_date, recent_value
    except Exception as e:
        print(f"Error fetching data: {e}")
        return None, None

def update_excel(file_path, recent_date, recent_value):
    """Update the Excel file with the new Existing Home Sales data and compute % change vs last year."""
    try:
        if not os.path.exists(file_path):
            print(f"File not found: {file_path}")
            return

        wb = load_workbook(file_path)
        sheet_name = "Data"
        if sheet_name not in wb.sheetnames:
            print(f"'{sheet_name}' sheet not found in workbook.")
            return

        ws = wb[sheet_name]

        # Read existing data from Excel
        existing_data = []
        for row in range(2, ws.max_row + 1):
            date_val = ws.cell(row=row, column=1).value
            value_val = ws.cell(row=row, column=2).value
            existing_data.append((date_val, value_val))

        # Check if the most recent date already exists
        if existing_data:
            most_recent_date_in_excel = existing_data[0][0]
            if pd.to_datetime(most_recent_date_in_excel).strftime("%Y-%m-%d") == recent_date:
                print("Data is already up to date. No changes made.")
                return

        # Insert new row at the top
        ws.insert_rows(2)
        ws.cell(row=2, column=1, value=recent_date)
        ws.cell(row=2, column=2, value=recent_value)
        existing_data.insert(0, (recent_date, recent_value))

        # Compute % change vs 12 months ago
        for i in range(len(existing_data)):
            row_excel = i + 2
            if i + 12 < len(existing_data):
                curr_val = existing_data[i][1]
                prev_val = existing_data[i + 12][1]
                if curr_val is not None and prev_val not in [0, None]:
                    change = (curr_val / prev_val - 1) * 100
                    ws.cell(row=row_excel, column=3, value=round(change, 1))
                else:
                    ws.cell(row=row_excel, column=3, value=None)
            else:
                ws.cell(row=row_excel, column=3, value=None)

        wb.save(file_path)
        print("✅ Excel file updated and saved successfully.")

    except Exception as e:
        print(f"Error updating Excel file: {e}")

if __name__ == "__main__":
    recent_date, recent_value = fetch_recent_sales_data(URL)
    if recent_date and recent_value:
        update_excel(EXCEL_FILE_PATH, recent_date, recent_value)

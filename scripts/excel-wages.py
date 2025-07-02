import os
import pandas as pd
from openpyxl import load_workbook
import requests
from bs4 import BeautifulSoup
from datetime import datetime

# File path to the Excel file
EXCEL_FILE_PATH = './data/wages.xlsx'

# YCharts URL for U.S. clothing store sales
URL = "https://ycharts.com/indicators/us_average_hourly_earnings"
HEADERS = {
    "User-Agent": "Mozilla/5.0",
    "Accept-Language": "en-US,en;q=0.9",
}

def fetch_recent_apparel_sales_data(url):
    """Fetch the most recent U.S. apparel store sales data from YCharts."""
    try:
        response = requests.get(url, headers=HEADERS)
        if response.status_code != 200:
            raise Exception(f"Request failed with status code {response.status_code}")

        soup = BeautifulSoup(response.content, "html.parser")
        element = soup.find("div", class_="key-stat-title")
        if not element:
            raise Exception("Failed to find the element with class 'key-stat-title'.")

        text = element.get_text(strip=True)
        if " USD for " not in text:
            raise Exception("Unexpected format: missing ' USD for '")

        value_str, date_str = text.split(" USD for ")
        value = float(value_str.replace(',', '').strip())
        date = datetime.strptime(date_str.strip(), "%b %Y").strftime("%Y-%m-%d")

        print(f"Fetched Apparel Sales - Value: {value}B, Date: {date}")
        return date, value
    except Exception as e:
        print(f"Error fetching apparel sales data: {e}")
        return None, None

def update_excel(file_path, recent_date, recent_value):
    """Update Excel with recent apparel sales data, and compute YoY % using Python."""
    try:
        if not os.path.exists(file_path):
            print(f"File not found: {file_path}")
            return

        wb = load_workbook(file_path)
        sheet_name = "Data"
        if sheet_name not in wb.sheetnames:
            print(f"'{sheet_name}' sheet not found.")
            return
        ws = wb[sheet_name]

        # Read existing data
        dates = []
        values = []

        for row in range(2, ws.max_row + 1):
            date_cell = ws.cell(row=row, column=1).value
            value_cell = ws.cell(row=row, column=2).value
            if date_cell and value_cell:
                dates.append(pd.to_datetime(date_cell))
                values.append(value_cell)

        if dates and pd.to_datetime(recent_date) == dates[0]:
            print(f"Date {recent_date} already exists. No update needed.")
            return

        # Insert new row at top
        ws.insert_rows(2)
        ws.cell(row=2, column=1, value=recent_date)
        ws.cell(row=2, column=2, value=recent_value)

        # Update lists
        dates.insert(0, pd.to_datetime(recent_date))
        values.insert(0, recent_value)

        # Compute YoY % using Python and write to column C
        for i in range(len(values)):
            if i + 12 < len(values):
                prev = values[i + 12]
                curr = values[i]
                if prev != 0:
                    yoy = ((curr / prev) - 1) * 100
                    ws.cell(row=i + 2, column=3, value=round(yoy, 2))
                else:
                    ws.cell(row=i + 2, column=3, value="N/A")
            else:
                ws.cell(row=i + 2, column=3, value="")

        wb.save(file_path)
        print(f"Excel updated with {recent_date} apparel data and YoY % computed.")
    except Exception as e:
        print(f"Error updating Excel file: {e}")

if __name__ == "__main__":
    date, value = fetch_recent_apparel_sales_data(URL)
    if date and value:
        update_excel(EXCEL_FILE_PATH, date, value)

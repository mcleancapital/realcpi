import os
import requests
import pandas as pd
from bs4 import BeautifulSoup
from datetime import datetime
from openpyxl import load_workbook

# Constants
EXCEL_FILE_PATH = "./data/sp-500-prices.xlsx"
SHEET_NAME = "Data"  # Ensure the sheet is always named "Data"
URL = "https://www.multpl.com/s-p-500-historical-prices/table/by-month"
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36",
    "Accept-Language": "en-US,en;q=0.9",
}

def fetch_latest_sp500_price(url):
    """Fetch the latest S&P 500 price from the webpage."""
    try:
        response = requests.get(url, headers=HEADERS)
        if response.status_code != 200:
            raise Exception(f"Failed to fetch data, Status Code: {response.status_code}")

        soup = BeautifulSoup(response.content, "html.parser")
        table_rows = soup.find_all("tr")

        if len(table_rows) < 2:
            raise Exception("No valid data found in the table.")

        latest_data = table_rows[1].find_all("td")  # Second row contains latest data
        if len(latest_data) < 2:
            raise Exception("Unexpected data format.")

        # Extract and clean the date
        latest_date = latest_data[0].text.strip()
        latest_value_str = latest_data[1].text.strip().replace("\u2020", "").replace("\n", "").strip()

        # Convert the value to a float
        latest_value = float(latest_value_str.replace(",", ""))

        # Convert date to YYYY-MM-DD format
        latest_date = datetime.strptime(latest_date, "%b %d, %Y").date()

        print(f"Fetched Latest Data - Date: {latest_date}, Value: {latest_value}")
        return latest_date, latest_value
    except Exception as e:
        print(f"Error fetching S&P 500 data: {e}")
        return None, None

def update_excel(file_path, latest_date, latest_value):
    """Update Excel file with new S&P 500 data, ensuring correct insertion rules."""
    try:
        # Load existing workbook or create a new one
        if os.path.exists(file_path):
            df = pd.read_excel(file_path, sheet_name=SHEET_NAME)
        else:
            df = pd.DataFrame(columns=["Date", "Value"])

        # Convert 'Date' column to datetime and normalize (removes time part)
        df["Date"] = pd.to_datetime(df["Date"], errors='coerce').dt.date

        # Check if there is existing data
        if not df.empty:
            most_recent_date = df.iloc[0]["Date"]  # Most recent row (top of DataFrame)

            # If the most recent date is NOT the first of the month, update it
            if most_recent_date.day != 1:
                print(f"Replacing row {most_recent_date} with {latest_date}.")
                df.at[0, "Date"] = latest_date
                df.at[0, "Value"] = latest_value
            else:
                # If the most recent row is the first of the month, add a new row
                new_row = pd.DataFrame([[latest_date, latest_value]], columns=["Date", "Value"])
                df = pd.concat([new_row, df], ignore_index=True)
                print(f"Added new data: {latest_date}, Value: {latest_value}")
        else:
            # If there is no data, initialize the file with the new row
            df = pd.DataFrame([[latest_date, latest_value]], columns=["Date", "Value"])
            print(f"Initializing data with {latest_date}, Value: {latest_value}")

        # Save the updated DataFrame to Excel
        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=SHEET_NAME, index=False)

        # Reopen Excel file to confirm data integrity
        wb = load_workbook(file_path)
        if SHEET_NAME not in wb.sheetnames:
            ws = wb.active
            ws.title = SHEET_NAME
        else:
            ws = wb[SHEET_NAME]

        # Save the final Excel file
        wb.save(file_path)
        print(f"Excel file updated successfully with new S&P 500 data in sheet '{SHEET_NAME}': {file_path}")

    except Exception as e:
        print(f"Error updating Excel file: {e}")

if __name__ == "__main__":
    # Fetch latest S&P 500 data
    latest_date, latest_value = fetch_latest_sp500_price(URL)

    # Update the Excel file if data was successfully fetched
    if latest_date and latest_value:
        update_excel(EXCEL_FILE_PATH, latest_date, latest_value)

import requests
import pandas as pd
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

# 🔐 Your API key from FMP
API_KEY = "y9Bthip8mNYaWhrHQp0eTtPX3KltVYPj"

# Index symbol for S&P/TSX Composite
INDEX_SYMBOL = "^GSPTSE"

# Excel file path
EXCEL_FILE = "./data/tsx.xlsx"
SHEET_NAME = "Data"

# Fetch monthly historical prices from FMP
def fetch_tsx_history(symbol, api_key):
    url = f"https://financialmodelingprep.com/api/v3/historical-price-full/index/{symbol}?timeseries=1000&apikey={api_key}"
    response = requests.get(url)
    response.raise_for_status()
    data = response.json()

    if "historical" not in data:
        raise Exception("No 'historical' key in response data")

    records = data["historical"]

    # Convert to DataFrame
    df = pd.DataFrame(records)
    df["date"] = pd.to_datetime(df["date"])
    df = df.sort_values("date", ascending=False).reset_index(drop=True)
    df = df[["date", "close"]].rename(columns={"date": "Date", "close": "Value"})

    # Calculate % change vs 252 rows later (approx. 1 year for daily data)
    pct_change = []
    for i in range(len(df)):
        if i + 252 < len(df):
            current = df.at[i, "Value"]
            past = df.at[i + 252, "Value"]
            if past != 0:
                change = ((current / past) - 1) * 100
                pct_change.append(round(change, 2))
            else:
                pct_change.append(None)
        else:
            pct_change.append(None)

    df["% Change vs Last Year"] = pct_change
    return df

# Write to Excel
def write_to_excel(df, file_path):
    if not os.path.exists(os.path.dirname(file_path)):
        os.makedirs(os.path.dirname(file_path))

    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        if SHEET_NAME in wb.sheetnames:
            del wb[SHEET_NAME]
    else:
        wb = Workbook()

    ws = wb.create_sheet(SHEET_NAME, 0)
    ws.append(["Date", "Value", "% Change vs Last Year"])

    for _, row in df.iterrows():
        ws.append([row["Date"].date(), row["Value"], row["% Change vs Last Year"]])

    wb.save(file_path)
    print(f"✔ Data saved to {file_path}")

# Run everything
try:
    print("🔍 Fetching TSX historical data from FMP...")
    tsx_df = fetch_tsx_history(INDEX_SYMBOL, API_KEY)
    print(f"✅ Fetched {len(tsx_df)} records.")
    write_to_excel(tsx_df, EXCEL_FILE)
except Exception as e:
    print(f"❌ Error: {e}")

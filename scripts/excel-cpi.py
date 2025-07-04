import os
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils.cell import get_column_letter
import requests
from bs4 import BeautifulSoup
from datetime import datetime

# File path to the Excel file
EXCEL_FILE_PATH = './data/cpi.xlsx'

# URL and headers
URL = "https://ycharts.com/indicators/us_consumer_price_index_yoy"
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36",
    "Accept-Language": "en-US,en;q=0.9",
}


def ensure_excel_exists(path):
    """Create Excel file and Data sheet if missing."""
    if not os.path.exists(path):
        print(f"{path} not found. Creating new file.")
        os.makedirs(os.path.dirname(path), exist_ok=True)
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        ws.append(["Date", "CPI", "YoY"])
        wb.save(path)
        print(f"Created file with headers: {path}")


def fetch_recent_cpi_data(url):
    """Fetch the most recent CPI data from the YCharts page."""
    try:
        response = requests.get(url, headers=HEADERS)
        if response.status_code != 200:
            raise Exception(f"Request failed with status code {response.status_code}")

        soup = BeautifulSoup(response.content, "html.parser")
        cpi_element = soup.find("div", class_="key-stat-title")
        if not cpi_element:
            raise Exception("Failed to find the element with class 'key-stat-title'.")

        cpi_text = cpi_element.get_text(strip=True)
        parts = cpi_text.split(maxsplit=2)
        if len(parts) < 2:
            raise Exception("Unexpected format for CPI data.")

        recent_value = float(parts[0].replace("%", ""))
        recent_date_str = parts[2].replace("for ", "")
        recent_date = datetime.strptime(recent_date_str, "%b %Y").strftime("%Y-%m-%d")

        print(f"Fetched CPI data - Value: {recent_value}, Date: {recent_date}")
        return recent_date, recent_value
    except Exception as e:
        print(f"Error fetching CPI data: {e}")
        return None, None


def update_excel(file_path, recent_date, recent_value):
    """Update the Excel file with the most recent CPI data."""
    try:
        wb = load_workbook(file_path)
        sheet_name = "Data"
        if sheet_name not in wb.sheetnames:
            print(f"'{sheet_name}' sheet not found. Creating it.")
            ws = wb.create_sheet(title=sheet_name)
            ws.append(["Date", "CPI", "YoY"])
        else:
            ws = wb[sheet_name]

        # Check if recent date already exists in A2
        most_recent_date_in_excel = ws.cell(row=2, column=1).value
        if most_recent_date_in_excel:
            most_recent_date_in_excel = pd.to_datetime(most_recent_date_in_excel).strftime("%Y-%m-%d")

        if most_recent_date_in_excel == recent_date:
            print(f"Recent date {recent_date} already exists. No update needed.")
            return

        # Insert a new row at the top (after headers)
        ws.insert_rows(2)
        ws.cell(row=2, column=1, value=recent_date)
        ws.cell(row=2, column=2, value=recent_value)

        # Update column C: static YoY values based on row+12 (if exists)
        for row in range(2, ws.max_row + 1):
            try:
                curr = ws.cell(row=row, column=2).value
                future = ws.cell(row=row + 12, column=2).value
                if isinstance(curr, (int, float)) and isinstance(future, (int, float)) and future != 0:
                    yoy = (curr / future - 1) * 100
                    ws.cell(row=row, column=3, value=round(yoy, 2))
                else:
                    ws.cell(row=row, column=3, value=None)
            except:
                ws.cell(row=row, column=3, value=None)

        wb.save(file_path)
        print(f"Excel file updated successfully: {file_path}")
    except Exception as e:
        print(f"Error updating Excel file: {e}")


if __name__ == "__main__":
    ensure_excel_exists(EXCEL_FILE_PATH)
    recent_date, recent_value = fetch_recent_cpi_data(URL)
    if recent_date and recent_value:
        update_excel(EXCEL_FILE_PATH, recent_date, recent_value)

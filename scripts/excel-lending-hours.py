import os
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import requests
from bs4 import BeautifulSoup

# === CONFIG ===
EXCEL_FILE_PATH = './data/lending-hours.xlsx'
SHEET_NAME = "Data"
URL = "https://ycharts.com/indicators/us_financial_activities_consumer_lending_aggregate_weekly_hours"
HEADERS = {
    "User-Agent": "Mozilla/5.0",
    "Accept-Language": "en-US,en;q=0.9",
}


def fetch_recent_data(url):
    """Fetch the most recent consumer lending hours data from YCharts."""
    try:
        response = requests.get(url, headers=HEADERS)
        if response.status_code != 200:
            raise Exception(f"Request failed with status code {response.status_code}")

        soup = BeautifulSoup(response.content, "html.parser")
        element = soup.find("div", class_="key-stat-title")
        if not element:
            raise Exception("Could not find 'key-stat-title' on the page.")

        text = element.get_text(strip=True)
        parts = text.split(maxsplit=2)
        if len(parts) < 3:
            raise Exception("Unexpected data format from YCharts.")

        value = float(parts[0].replace('M', ''))
        date = datetime.strptime(parts[2].replace("for ", ""), "%b %Y").strftime("%Y-%m-%d")

        print(f"✔️  Fetched value: {value}M for {date}")
        return date, value
    except Exception as e:
        print(f"❌ Error fetching data: {e}")
        return None, None


def update_excel(file_path, new_date, new_value):
    """Update Excel with new data and compute % change vs last year using pandas."""
    try:
        if not os.path.exists(file_path):
            print(f"❌ File not found: {file_path}")
            return

        # Load existing data into DataFrame
        wb = load_workbook(file_path)
        if SHEET_NAME not in wb.sheetnames:
            print(f"❌ Sheet '{SHEET_NAME}' not found.")
            return
        ws = wb[SHEET_NAME]

        data = []
        for row in range(2, ws.max_row + 1):
            dt = ws.cell(row=row, column=1).value
            val = ws.cell(row=row, column=2).value
            if dt and val:
                data.append((pd.to_datetime(dt), float(val)))

        # Check if date already exists
        if data and data[0][0].strftime("%Y-%m-%d") == new_date:
            print(f"ℹ️  {new_date} already exists — skipping insert.")
            return

        # Insert new row at top
        data.insert(0, (pd.to_datetime(new_date), float(new_value)))

        # Compute % change vs 12 months ago
        values = [v for (_, v) in data]
        changes = []
        for i in range(len(values)):
            if i + 12 < len(values) and values[i + 12] != 0:
                pct = (values[i] / values[i + 12] - 1) * 100
                changes.append(round(pct, 1))
            else:
                changes.append(None)

        # Write back to Excel
        for i, (dt, val) in enumerate(data):
            ws.cell(row=i + 2, column=1, value=dt.strftime("%Y-%m-%d"))
            ws.cell(row=i + 2, column=2, value=val)
            ws.cell(row=i + 2, column=3, value=changes[i])

        wb.save(file_path)
        print("✅ Excel updated with static values — no formulas required.")

    except Exception as e:
        print(f"❌ Error updating Excel: {e}")


if __name__ == "__main__":
    date, value = fetch_recent_data(URL)
    if date and value:
        update_excel(EXCEL_FILE_PATH, date, value)

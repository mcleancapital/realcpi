import os
import pandas as pd
from openpyxl import load_workbook
import requests
from bs4 import BeautifulSoup
from datetime import datetime

# File path to the Excel file
EXCEL_FILE_PATH = './data/lei.xlsx'

# URL and headers
URL = "https://tradingeconomics.com/united-states/leading-economic-index"
HEADERS = {
    "User-Agent": "Mozilla/5.0",
    "Accept-Language": "en-US,en;q=0.9",
}

def fetch_recent_lei_data(url):
    """Fetch the most recent LEI value and date from TradingEconomics."""
    try:
        response = requests.get(url, headers=HEADERS)
        if response.status_code != 200:
            raise Exception(f"Request failed with status code {response.status_code}")

        soup = BeautifulSoup(response.content, "html.parser")

        # Find all <td> elements
        td_elements = soup.find_all("td")

        # Extract the first numeric value for LEI
        recent_value = None
        for td in td_elements:
            text = td.get_text(strip=True)
            if text.replace('.', '', 1).isdigit():
                recent_value = float(text)
                break

        if recent_value is None:
            raise Exception("No numeric value found in <td> elements.")

        # Extract the date from the 5th <td> element
        if len(td_elements) < 5:
            raise Exception("Less than 5 <td> elements found.")
        recent_date_str = td_elements[4].get_text(strip=True)

        # Parse date in "MMM YYYY" format
        recent_date = datetime.strptime(recent_date_str, "%b %Y").replace(day=1).strftime("%Y-%m-%d")

        print(f"Fetched LEI data - Value: {recent_value}, Date: {recent_date}")
        return recent_date, recent_value
    except Exception as e:
        print(f"Error fetching LEI data: {e}")
        return None, None

def update_excel(file_path, recent_date, recent_value):
    """Update the Excel file with the most recent LEI data and compute YoY % in column C."""
    try:
        if not os.path.exists(file_path):
            print(f"File not found: {file_path}")
            return

        wb = load_workbook(file_path)
        sheet_name = "Data"
        if sheet_name not in wb.sheetnames:
            print(f"'{sheet_name}' sheet not found in the workbook.")
            return
        ws = wb[sheet_name]

        # Check most recent existing date in Excel
        most_recent_date_in_excel = ws.cell(row=2, column=1).value
        if most_recent_date_in_excel:
            most_recent_date_in_excel = pd.to_datetime(most_recent_date_in_excel).strftime("%Y-%m-%d")
        if most_recent_date_in_excel == recent_date:
            print(f"Recent date {recent_date} already exists in the Excel file. No update needed.")
            return

        # Insert new row at the top
        ws.insert_rows(2)
        ws.cell(row=2, column=1, value=recent_date)
        ws.cell(row=2, column=2, value=recent_value)

        # Gather all LEI values from column B
        values = []
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row=row, column=2).value
            values.append(val)

        # Compute YoY % in column C using Python
        for i in range(len(values)):
            if i + 12 < len(values):
                prev = values[i + 12]
                curr = values[i]
                if prev and prev != 0:
                    yoy = ((curr / prev) - 1) * 100
                    ws.cell(row=i + 2, column=3, value=round(yoy, 2))
                else:
                    ws.cell(row=i + 2, column=3, value="N/A")
            else:
                ws.cell(row=i + 2, column=3, value="")

        wb.save(file_path)
        print(f"Excel file updated successfully with LEI data and Python-computed YoY %.")
    except Exception as e:
        print(f"Error updating Excel file: {e}")

if __name__ == "__main__":
    recent_date, recent_value = fetch_recent_lei_data(URL)
    if recent_date and recent_value:
        update_excel(EXCEL_FILE_PATH, recent_date, recent_value)

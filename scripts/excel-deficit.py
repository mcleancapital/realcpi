import requests
from openpyxl import Workbook, load_workbook
from datetime import datetime

# FRED API Setup
API_KEY = "dc0a72da11fb0ffc6f4dc8d31e7afbb5"  # Replace with your FRED API key
FRED_BASE = "https://api.stlouisfed.org/fred/series/observations"
SERIES_ID = "FYFSGDA188S"  # Federal Surplus or Deficit as % of GDP

# Output Excel file
EXCEL_FILE = "./data/us-deficit.xlsx"

# Fetch data from FRED
def get_fred_annual_data(series_id, api_key):
    params = {
        "series_id": series_id,
        "api_key": api_key,
        "file_type": "json",
        "frequency": "a"  # Annual frequency
    }
    response = requests.get(FRED_BASE, params=params)
    response.raise_for_status()
    data = response.json()
    if "observations" in data:
        return {
            f"{obs['date'][:4]}-12-31": float(obs["value"])
            for obs in data["observations"]
            if obs["value"] != "."
        }
    else:
        raise Exception(f"No data found for series {series_id}")

# Main logic
try:
    deficit_data = get_fred_annual_data(SERIES_ID, API_KEY)

    # Get the most recent year
    latest_year = max(deficit_data.keys())
    latest_value = deficit_data[latest_year]

    # Load or create Excel file
    try:
        wb = load_workbook(EXCEL_FILE)
        sheet = wb.active
    except FileNotFoundError:
        wb = Workbook()
        sheet = wb.active
        sheet.append(["Year", "Deficit/Surplus (% of GDP)"])

    # Get existing years to avoid duplicates
    existing_years = [str(sheet.cell(row=i, column=1).value) for i in range(2, sheet.max_row + 1)]
    if latest_year not in existing_years:
        sheet.insert_rows(2)
        sheet.cell(row=2, column=1, value=latest_year)
        sheet.cell(row=2, column=2, value=latest_value)

    # Save
    wb.save(EXCEL_FILE)
    print(f"{latest_year}: {latest_value:.2f}% saved to {EXCEL_FILE}")

except Exception as e:
    print(f"Error: {e}")

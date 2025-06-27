import os
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import requests
from bs4 import BeautifulSoup

# File path to the Excel file
EXCEL_FILE_PATH = './data/pmi.xlsx'

# URL and headers
URL = "https://ycharts.com/indicators/us_pmi"
HEADERS = {
    "User-Agent": "Mozilla/5.0",
    "Accept-Language": "en-US,en;q=0.9",
}

def fetch_recent_pmi_data(url):
    """Fetch the most recent PMI data from YCharts."""
    try:
        response = requests.get(url, headers=HEADERS)
        if response.status_code != 200:
            raise Exception(f"Request failed with status code {response.status_code}")

        soup = BeautifulSoup(response.content, "html.parser")
        element = soup.find("div", class_="key-stat-title")
        if not element:
            raise Exception("Could not find 'key-stat-title'")

        text = element.get_text(strip=True)
        parts = text.split(maxsplit=2)
        if len(parts) < 3:
            raise Exception("Unexpected PMI format")

        value = float(parts[0])
        date_str = parts[2].replace("for ", "")
        date = datetime.strptime(date_str, "%b %Y").strftime("%Y-%m-%d")

        print(f"Fetched PMI: {value} on {date}")
        return date, value
    except Exception as e:
        print(f"Error fetching PMI data: {e}")
        return None, None

def update_excel(file_path, new_date, new_value):
    try:
        if not os.path.exists(file_path):
            print(f"File not found: {file_path}")
            return

        wb = load_workbook(file_path)
        sheet_name = "Data"
        if sheet_name not in wb.sheetnames:
            print(f"'{sheet_name}' sheet not found in workbook.")
            return

        ws = wb[sheet_name]

        # Read existing rows
        data = []
        for row in range(2, ws.max_row + 1):
            date_val = ws.cell(row=row, column=1).value
            value_val = ws.cell(row=row, column=2).value
            if date_val and value_val:
                data.append((pd.to_datetime(date_val).strftime("%Y-%m-%d"), float(value_val)))

        # Check if latest already exists
        all_dates = {d[0] for d in data}
        if new_date in all_dates:
            print(f"Date {new_date} already in Excel. No update needed.")
            return

        # Add new data to top
        data.insert(0, (new_date, new_value))

        # Write new top row
        ws.insert_rows(2)
        ws.cell(row=2, column=1, value=new_date)
        ws.cell(row=2, column=2, value=new_value)

        # Recompute % change vs 12 months ago
        for i in range(len(data)):
            row_excel = i + 2
            if i + 12 < len(data):
                curr = data[i][1]
                prev = data[i + 12][1]
                if prev != 0:
                    change = (curr / prev - 1) * 100
                    ws.cell(row=row_excel, column=3, value=round(change, 1))
                else:
                    ws.cell(row=row_excel, column=3, value=None)
            else:
                ws.cell(row=row_excel, column=3, value=None)

        wb.save(file_path)
        print(f"✅ Excel updated with PMI for {new_date}")
    except Exception as e:
        print(f"❌ Error updating Excel file: {e}")

if __name__ == "__main__":
    date, value = fetch_recent_pmi_data(URL)
    if date and value:
        update_excel(EXCEL_FILE_PATH, date, value)

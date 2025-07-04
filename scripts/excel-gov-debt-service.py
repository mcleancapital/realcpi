import requests
from openpyxl import Workbook, load_workbook
from datetime import datetime

# FRED API Setup
API_KEY = "dc0a72da11fb0ffc6f4dc8d31e7afbb5"  # Replace with your own key
FRED_BASE = "https://api.stlouisfed.org/fred/series/observations"
SERIES_ID = "FYOIGDA188S"  # Interest Payments on Federal Debt as a % of GDP

# Output Excel file
EXCEL_FILE = "./data/gov-debt-service.xlsx"

# Fetch annual data from FRED
def get_fred_annual_data(series_id, api_key):
    params = {
        "series_id": series_id,
        "api_key": api_key,
        "file_type": "json",
        "frequency": "a"  # Annual frequency
    }
    response = requests.get(FRED_BASE, params=params)
    response.raise_for_status()
    data = response.json()
    if "observations" in data:
        return {
            obs["date"]: float(obs["value"])
            for obs in data["observations"]
            if obs["value"] != "."
        }
    else:
        raise Exception(f"No data found for series {series_id}")

# Main logic
try:
    # Step 1: Get latest data
    data = get_fred_annual_data(SERIES_ID, API_KEY)
    latest_date = max(data.keys())
    latest_value = data[latest_date]

    # Step 2: Load or create Excel file
    try:
        wb = load_workbook(EXCEL_FILE)
        sheet = wb.active
    except FileNotFoundError:
        wb = Workbook()
        sheet = wb.active
        sheet.append(["Date", "Value", "% Change vs Last Year"])

    # Step 3: Check if latest date is already in the file
    existing_dates = [str(sheet.cell(row=i, column=1).value) for i in range(2, sheet.max_row + 1)]

    if latest_date not in existing_dates:
        print(f"Inserting new row for {latest_date}...")

        # Insert new row at the top (row 2)
        sheet.insert_rows(2)
        sheet.cell(row=2, column=1, value=latest_date)
        sheet.cell(row=2, column=2, value=latest_value)

        # Compute % change from previous year (row 3)
        try:
            prev_value = sheet.cell(row=3, column=2).value
            if prev_value not in (None, 0):
                change = (latest_value / float(prev_value) - 1) * 100
                sheet.cell(row=2, column=3, value=round(change, 2))
            else:
                sheet.cell(row=2, column=3, value="")
        except Exception:
            sheet.cell(row=2, column=3, value="")
    else:
        print(f"{latest_date} already exists — skipping insertion.")

    # Step 4: Recalculate column C for all rows
    for row in range(2, sheet.max_row):
        try:
            current_val = sheet.cell(row=row, column=2).value
            next_val = sheet.cell(row=row + 1, column=2).value
            if current_val not in (None, "") and next_val not in (None, "", 0):
                change = (float(current_val) / float(next_val) - 1) * 100
                sheet.cell(row=row, column=3, value=round(change, 2))
            else:
                sheet.cell(row=row, column=3, value="")
        except Exception:
            sheet.cell(row=row, column=3, value="")

    # Step 5: Save workbook
    wb.save(EXCEL_FILE)
    print(f"{latest_date}: {latest_value:.2f}% saved to {EXCEL_FILE}")

except Exception as e:
    print(f"Error: {e}")

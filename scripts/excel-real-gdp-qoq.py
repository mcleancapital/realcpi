import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter
import requests
from bs4 import BeautifulSoup
from datetime import datetime

# File path to the Excel file
EXCEL_FILE_PATH = './data/us-real-gdp-growth-qoq.xlsx'

# URL and headers
URL = "https://ycharts.com/indicators/us_real_gdp_growth"
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36",
    "Accept-Language": "en-US,en;q=0.9",
}

def convert_quarter_to_date(quarter_str):
    """Convert 'Q1 2024' format to the last month of the quarter ('2024-03-31' for Q1)."""
    quarter_map = {
        "Q1": "03-31",  # March
        "Q2": "06-30",  # June
        "Q3": "09-30",  # September
        "Q4": "12-31"   # December
    }
    
    parts = quarter_str.split()
    if len(parts) != 2 or parts[0] not in quarter_map or not parts[1].isdigit():
        raise ValueError(f"Unexpected quarter format: {quarter_str}")

    quarter, year = parts[0], parts[1]
    return f"{year}-{quarter_map[quarter]}"

def fetch_recent_cpi_data(url):
    """Fetch the most recent GDP growth rate from the YCharts page."""
    try:
        response = requests.get(url, headers=HEADERS)
        if response.status_code != 200:
            raise Exception(f"Request failed with status code {response.status_code}")

        soup = BeautifulSoup(response.content, "html.parser")
        cpi_element = soup.find("div", class_="key-stat-title")
        if not cpi_element:
            raise Exception("Failed to find the element with class 'key-stat-title'.")

        cpi_text = cpi_element.get_text(strip=True)
        parts = cpi_text.split(maxsplit=2)
        if len(parts) < 2:
            raise Exception("Unexpected format for GDP Growth data.")

        recent_value = float(parts[0].replace('%', ''))  # Extract GDP Growth Rate
        recent_date = convert_quarter_to_date(parts[2].replace("for ", ""))  # Convert to 'YYYY-MM-DD'

        print(f"Fetched GDP Growth data - Value: {recent_value}, Date: {recent_date}")
        return recent_date, recent_value
    except Exception as e:
        print(f"Error fetching GDP Growth data: {e}")
        return None, None

def update_excel(file_path, recent_date, recent_value):
    """Update the Excel file with the most recent GDP Growth data."""
    try:
        # Check if the file exists
        if not os.path.exists(file_path):
            print(f"File not found: {file_path}")
            return

        # Load the workbook
        wb = load_workbook(file_path)
        sheet_name = "Data"  # Assuming your data is in a sheet named 'Data'
        if sheet_name not in wb.sheetnames:
            print(f"'{sheet_name}' sheet not found in the workbook.")
            return
        ws = wb[sheet_name]

        # Check the most recent date in A2
        most_recent_date_in_excel = ws.cell(row=2, column=1).value
        if most_recent_date_in_excel:
            most_recent_date_in_excel = pd.to_datetime(most_recent_date_in_excel).strftime("%Y-%m-%d")

        # Add a new row only if the date is not already present
        if most_recent_date_in_excel == recent_date:
            print(f"Recent date {recent_date} already exists in the Excel file. No update needed.")
            return

        # Insert a new row
        ws.insert_rows(2)
        ws.cell(row=2, column=1, value=recent_date)  # Date (YYYY-MM-DD)
        ws.cell(row=2, column=2, value=recent_value)  # US Real GDP Growth Rate

        # Update formulas in column C for all rows
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=3, value=f"=(B{row}/B{row+4}-1)*100")  # Adjusting formula for quarterly data

        print(f"Added new data - Date: {recent_date}, Value: {recent_value}, Updated formulas in column C.")

        # Save the workbook
        wb.save(file_path)
        print(f"Excel file updated successfully: {file_path}")
    except Exception as e:
        print(f"Error updating Excel file: {e}")

if __name__ == "__main__":
    # Fetch the most recent GDP Growth data
    recent_date, recent_value = fetch_recent_cpi_data(URL)

    # Update the Excel file if data is successfully fetched
    if recent_date and recent_value:
        update_excel(EXCEL_FILE_PATH, recent_date, recent_value)
